#!/usr/bin/env python3
"""
Invoice Generator v0.2 — GST Tax Invoice from Raw Delivery Data.

Reads structured delivery/challan records from an Excel file and produces
professionally formatted, GST-compliant tax invoices (Excel + PDF).

CLI Usage:
    python3 generate_invoices.py                               # Interactive
    python3 generate_invoices.py -i 178                        # Invoice #178
    python3 generate_invoices.py -i 178 --pdf                  # Also generate PDF
    python3 generate_invoices.py --batch ./data/ --start 178   # Batch process

Module Usage:
    from generate_invoices import generate_invoice, generate_pdf
    wb = generate_invoice("data.xlsx", inv_num=178)
    wb.save("output.xlsx")
    generate_pdf("data.xlsx", inv_num=178, output_path="output.pdf")
"""

__version__ = "0.2.0"

import argparse
import calendar
import math
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Optional, Union

import pandas as pd
import yaml
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side


# ─── Layout Constants ────────────────────────────────────────────────────────
# These control the physical appearance of the Excel invoice.
# Heights are in Excel "points" (roughly 1pt ≈ 0.35mm).

COLUMN_WIDTHS = {
    "A": 11.57,   # Date
    "B": 10.29,   # Material
    "C":  7.29,   # Site
    "D": 10.29,   # Challan
    "E": 11.29,   # Quantity
    "F": 24.14,   # Rate / Tax label
    "G": 15.43,   # Per / Tax %
    "H": 10.57,   # Amount
}
TOTAL_COLUMNS = 8

ROW_HEIGHT_COMPANY_NAME = 58.5
ROW_HEIGHT_DEFAULT = 13.5
ROW_HEIGHT_BUYER_INFO = 51.0
ROW_HEIGHT_BANK_DETAILS = 57.75
ROW_HEIGHT_TERMS_SECOND = 12.0
ROW_HEIGHT_TERMS_THIRD = 21.75

# The spacer row fills empty space so the invoice occupies one printed page.
# This base height is calibrated for 9 material rows (3 sites × 3 materials).
SPACER_BASE_HEIGHT = 304.5
SPACER_BASELINE_ROWS = 9
SPACER_MIN_HEIGHT = 10.0

# First row where material data begins
MATERIAL_START_ROW = 10

# Date formats we attempt when parsing string dates, in priority order
SUPPORTED_DATE_FORMATS = [
    "%d/%m/%Y",    # 16/02/2025 (most common in India)
    "%d-%m-%Y",    # 16-02-2025
    "%Y-%m-%d",    # 2025-02-16 (ISO)
    "%m/%d/%Y",    # 02/16/2025 (US)
    "%d/%m/%y",    # 16/02/25
]

# Excel sheet names: max 31 chars, these characters are forbidden
_INVALID_SHEET_CHARS = re.compile(r'[\\/*?\[\]:]')
_MAX_SHEET_NAME_LENGTH = 31

# Required columns in the input Excel file
REQUIRED_COLUMNS = frozenset({
    "Date", "Challan No.", "Vehicle No.", "Site",
    "Material", "Quantity", "Rate", "Per",
})

# Currency format used throughout the invoice
CURRENCY_FORMAT = '[$₹]#,##0.00'
PERCENTAGE_FORMAT = '0.00%'
NUMBER_FORMAT = '#,##0.00'


# ─── Border Styles ───────────────────────────────────────────────────────────
# Named for clarity instead of cryptic abbreviations.

BORDER_ALL = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
BORDER_SIDES = Border(
    left=Side(style="thin"), right=Side(style="thin"),
)
BORDER_SIDES_TOP = Border(
    left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"),
)
BORDER_SIDES_BOTTOM = Border(
    left=Side(style="thin"), right=Side(style="thin"), bottom=Side(style="thin"),
)


# ─── Configuration ───────────────────────────────────────────────────────────

INPUT_FILE = Path("Example Data/Raw_Data.xlsx")
OUTPUT_EXCEL = Path("Final_Output.xlsx")
CONFIG_FILE = Path("config.yaml")

DEFAULT_CONFIG = {
    "company": {
        "name": "YOUR COMPANY NAME",
        "subtitle": "(BUSINESS TYPE)",
        "address": "Your Address Here",
        "contact": "0000000000",
        "gstn": "00XXXXX0000X0XX",
        "pan": "XXXXX0000X",
    },
    "buyer": {
        "name": "BUYER NAME",
        "address": "Buyer Address",
        "gstn": "00XXXXX0000X0XX",
    },
    "bank": {
        "account_name": "YOUR COMPANY NAME",
        "bank_name": "BANK NAME",
        "account_no": "000000000000",
        "branch": "BRANCH",
        "ifsc": "XXXX0000000",
    },
    "gst": {
        "cgst_rate": 0.09,
        "sgst_rate": 0.09,
        "hsn_code": 996511,
    },
    "unit": "Tonne",
}


def load_config(config_path: Optional[Union[str, Path]] = None) -> dict:
    """
    Load business configuration from a YAML file.

    If the file doesn't exist, returns sensible defaults. User-provided
    values are merged on top of defaults, so partial configs work fine.

    Args:
        config_path: Path to YAML config. Defaults to ./config.yaml.

    Returns:
        Complete configuration dict with keys: company, buyer, bank, gst, unit.
    """
    path = Path(config_path) if config_path else CONFIG_FILE

    if not path.exists():
        return DEFAULT_CONFIG.copy()

    with open(path, "r", encoding="utf-8") as file:
        user_config = yaml.safe_load(file) or {}

    # Deep-merge: user values override defaults, but missing keys get defaults
    merged = {}
    for section_name, default_values in DEFAULT_CONFIG.items():
        if isinstance(default_values, dict):
            merged[section_name] = {**default_values, **user_config.get(section_name, {})}
        else:
            merged[section_name] = user_config.get(section_name, default_values)

    return merged


# ─── Utility Functions ───────────────────────────────────────────────────────

def sanitize_sheet_name(name: str) -> str:
    """
    Make a string safe for use as an Excel sheet name.

    Excel sheet names cannot contain \\ / * ? [ ] : and must be ≤31 characters.
    Invalid characters are replaced with underscores.
    """
    cleaned = _INVALID_SHEET_CHARS.sub("_", str(name))
    return cleaned[:_MAX_SHEET_NAME_LENGTH]


def get_fiscal_year(dates: list[datetime]) -> str:
    """
    Determine the Indian fiscal year from a list of dates.

    The Indian fiscal year runs April–March. A date in Feb 2025 belongs
    to FY 2024/25, while a date in May 2025 belongs to FY 2025/26.

    Returns:
        String like '2024/25' or '2025/26'.
    """
    latest_date = max(dates)
    start_year = latest_date.year if latest_date.month >= 4 else latest_date.year - 1
    end_year_suffix = str(start_year + 1)[-2:]
    return f"{start_year}/{end_year_suffix}"


def get_invoice_date(dates: list[datetime]) -> datetime:
    """
    Compute the invoice date as the last day of the latest month in the data.

    For example, if the latest delivery is Feb 20, the invoice date is Feb 28.
    """
    latest_date = max(dates)
    last_day_of_month = calendar.monthrange(latest_date.year, latest_date.month)[1]
    return datetime(latest_date.year, latest_date.month, last_day_of_month)


def format_date_short(date: datetime) -> str:
    """
    Format a date as 'D-Mon' (e.g., '5-Feb', '16-Mar').

    Uses f-string instead of strftime's %-d flag for cross-platform safety
    (%-d works on macOS/Linux but not Windows).
    """
    return f"{date.day}-{date.strftime('%b')}"


def parse_dates(date_series: pd.Series) -> list[datetime]:
    """
    Parse a column of date values into datetime objects.

    Handles multiple input types gracefully:
    - datetime objects (passed through)
    - pandas Timestamps (converted to datetime)
    - Strings in various formats (tried against SUPPORTED_DATE_FORMATS)
    - NaN/None values (skipped)
    - Unparseable strings (skipped without crashing)

    Returns:
        List of successfully parsed datetime objects. May be empty.
    """
    parsed = []

    for value in date_series:
        if pd.isna(value):
            continue

        if isinstance(value, pd.Timestamp):
            parsed.append(value.to_pydatetime())
        elif isinstance(value, datetime):
            parsed.append(value)
        else:
            # Try each known format until one works
            date_string = str(value).strip()
            for date_format in SUPPORTED_DATE_FORMATS:
                try:
                    parsed.append(datetime.strptime(date_string, date_format))
                    break
                except ValueError:
                    continue
            # If nothing matched, silently skip this value

    return parsed


def round_conventional(value: float) -> int:
    """
    Round to nearest integer with conventional rounding (0.5 rounds up).

    Python's built-in round() uses "banker's rounding" (round-half-to-even),
    which gives unexpected results for business: round(100.5) → 100.
    This function gives the expected: round_conventional(100.5) → 101.
    """
    return int(math.floor(value + 0.5))


# ─── Data Processing ─────────────────────────────────────────────────────────

def process_input_data(
    input_file: Union[str, Path],
    invoice_number: int,
    config: Optional[dict] = None,
) -> dict[str, Any]:
    """
    Read, validate, and structure raw delivery data for invoice generation.

    This is the shared data pipeline used by both Excel and PDF generators.
    It handles all validation, cleaning, and computation so the rendering
    functions can focus purely on formatting.

    Args:
        input_file: Path to the input Excel file (.xlsx or .xls).
        invoice_number: The invoice number to assign.
        config: Business configuration dict. If None, loads from config.yaml.

    Returns:
        Dictionary containing all processed data needed for rendering:
        - df: Cleaned DataFrame
        - sites_data: Per-site material breakdowns
        - materials: Ordered list of material names
        - fiscal_year, invoice_date: Derived date info
        - total_amount, cgst_amount, sgst_amount, round_off, grand_total

    Raises:
        FileNotFoundError: If the input file doesn't exist.
        ValueError: If the data is missing columns, has no valid rows,
                    contains non-numeric quantities/rates, or has no parseable dates.
    """
    if config is None:
        config = load_config()

    # ── Validate file exists ──
    if isinstance(input_file, (str, Path)):
        file_path = Path(input_file)
        if not file_path.exists():
            raise FileNotFoundError(f"Input file not found: {file_path}")

    # ── Read Excel, auto-detecting format ──
    engine = "openpyxl"
    if isinstance(input_file, (str, Path)) and Path(input_file).suffix.lower() == ".xls":
        engine = "xlrd"

    dataframe = pd.read_excel(input_file, engine=engine)

    # ── Validate required columns exist ──
    missing_columns = REQUIRED_COLUMNS - set(dataframe.columns)
    if missing_columns:
        raise ValueError(
            f"Missing required columns: {', '.join(sorted(missing_columns))}"
        )

    # ── Drop summary/total rows (identified by NaN in Site column) ──
    dataframe = dataframe.dropna(subset=["Site"])
    if dataframe.empty:
        raise ValueError("No valid data rows found (all rows have empty 'Site').")

    # ── Coerce numeric columns ──
    # This catches text like "N/A" or "-" in Quantity/Rate and converts them to NaN
    for column_name in ["Quantity", "Rate"]:
        dataframe[column_name] = pd.to_numeric(dataframe[column_name], errors="coerce")

    invalid_rows = dataframe[dataframe["Quantity"].isna() | dataframe["Rate"].isna()]
    if not invalid_rows.empty:
        # Report Excel row numbers (1-indexed header + 1-indexed data)
        row_numbers = ", ".join(str(index + 2) for index in invalid_rows.index[:10])
        raise ValueError(
            f"Non-numeric or empty values in Quantity/Rate at Excel rows: {row_numbers}"
        )

    # ── Warn about negative quantities (returns/credits) ──
    negative_rows = dataframe[dataframe["Quantity"] < 0]
    if not negative_rows.empty:
        row_numbers = ", ".join(str(index + 2) for index in negative_rows.index[:5])
        print(
            f"⚠️  Warning: {len(negative_rows)} row(s) have negative quantities "
            f"(rows: {row_numbers}). These may represent returns/credits.",
            file=sys.stderr,
        )

    # ── Parse and validate dates ──
    all_dates = parse_dates(dataframe["Date"])
    if not all_dates:
        raise ValueError("No valid dates found in the 'Date' column.")

    fiscal_year = get_fiscal_year(all_dates)
    invoice_date = get_invoice_date(all_dates)

    # ── Discover unique sites and materials (preserving first-appearance order) ──
    sites = list(dict.fromkeys(dataframe["Site"]))
    materials = list(dict.fromkeys(dataframe["Material"]))
    unit = config.get("unit", "Tonne")

    # ── Build per-site breakdowns ──
    sites_data = []
    for site_name in sites:
        site_rows = dataframe[dataframe["Site"] == site_name]
        site_dates = parse_dates(site_rows["Date"])
        earliest_date = min(site_dates)
        latest_date = max(site_dates)

        # Show "16-Feb" for single-day, "16-Feb\nto\n28-Feb" for ranges
        if earliest_date.date() == latest_date.date():
            date_range_label = format_date_short(earliest_date)
        else:
            date_range_label = (
                f"{format_date_short(earliest_date)}\n"
                f"to\n"
                f"{format_date_short(latest_date)}"
            )

        # Aggregate each material's quantity and capture its rate
        material_data = {}
        for material_name in materials:
            material_rows = site_rows[site_rows["Material"] == material_name]
            if len(material_rows) > 0:
                material_data[material_name] = {
                    "qty": round(material_rows["Quantity"].sum(), 2),
                    "rate": float(material_rows["Rate"].iloc[0]),
                }

        sites_data.append({
            "site_name": site_name,
            "date_range_label": date_range_label,
            "material_data": material_data,
        })

    # ── Compute invoice totals ──
    cgst_rate = config["gst"]["cgst_rate"]
    sgst_rate = config["gst"]["sgst_rate"]

    total_amount = round(sum(
        material_info.get("qty", 0) * material_info.get("rate", 0)
        for site in sites_data
        for material_info in site["material_data"].values()
    ), 2)

    cgst_amount = round(total_amount * cgst_rate, 2)
    sgst_amount = round(total_amount * sgst_rate, 2)
    gross_total = total_amount + cgst_amount + sgst_amount
    grand_total = round_conventional(gross_total)
    round_off = round(grand_total - gross_total, 2)

    return {
        "df": dataframe,
        "sites_data": sites_data,
        "materials": materials,
        "fiscal_year": fiscal_year,
        "invoice_date": invoice_date,
        "inv_num": invoice_number,
        "config": config,
        "unit": unit,
        "total_amount": total_amount,
        "cgst_amount": cgst_amount,
        "sgst_amount": sgst_amount,
        "round_off": round_off,
        "grand_total": grand_total,
    }


# ─── Excel Formatting Helpers ────────────────────────────────────────────────

def style_cell(
    cell,
    font_name: str = "Arial",
    font_size: int = 9,
    bold: bool = False,
    horizontal: str = "center",
    vertical: str = "center",
    wrap: bool = True,
    border: Optional[Border] = None,
    number_format: Optional[str] = None,
) -> None:
    """Apply font, alignment, border, and number format to an Excel cell."""
    cell.font = Font(name=font_name, size=font_size, bold=bold)
    cell.alignment = Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format


def _apply_merged_row_borders(worksheet, row: int, total_cols: int = TOTAL_COLUMNS) -> None:
    """Apply proper borders to a merged row spanning all columns."""
    worksheet.cell(row=row, column=1).border = Border(
        left=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for column in range(2, total_cols):
        worksheet.cell(row=row, column=column).border = Border(
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
    worksheet.cell(row=row, column=total_cols).border = Border(
        right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"),
    )


def _apply_side_borders(worksheet, row: int, start_col: int = 1, end_col: int = TOTAL_COLUMNS,
                         border: Border = BORDER_SIDES) -> None:
    """Apply the same border to a range of columns in a row."""
    for column in range(start_col, end_col + 1):
        worksheet.cell(row=row, column=column).border = border


def _write_tax_line(worksheet, row: int, label: str, rate: Optional[float],
                    amount: float, border: Border = BORDER_SIDES) -> None:
    """Write a tax summary row (TOTAL, CGST, SGST) with consistent formatting."""
    _apply_side_borders(worksheet, row, end_col=TOTAL_COLUMNS - 1)

    cell = worksheet.cell(row=row, column=5, value=label)
    style_cell(cell, bold=(label == "TOTAL"), border=border)

    if rate is not None:
        cell = worksheet.cell(row=row, column=6, value=rate)
        style_cell(cell, font_size=10, number_format=PERCENTAGE_FORMAT, border=border)

    cell = worksheet.cell(row=row, column=8, value=amount)
    style_cell(cell, font_size=10, bold=(label == "TOTAL"),
               number_format=CURRENCY_FORMAT, border=border)


# ─── Excel Sheet Builders ────────────────────────────────────────────────────

def _create_invoice_sheet(workbook: Workbook, data: dict) -> None:
    """
    Create the formatted Bill (tax invoice) sheet.

    This is the primary output: a professional invoice with company header,
    buyer details, itemized materials per site, tax calculations, bank
    details, and terms & conditions.
    """
    invoice_number = data["inv_num"]
    config = data["config"]
    sites_data = data["sites_data"]
    materials = data["materials"]
    unit = data["unit"]
    fiscal_year = data["fiscal_year"]
    invoice_date = data["invoice_date"]

    sheet_name = sanitize_sheet_name(f"Bill - {invoice_number}")
    worksheet = workbook.create_sheet(title=sheet_name)

    num_material_rows = len(sites_data) * len(materials)

    # ── Column widths ──
    for letter, width in COLUMN_WIDTHS.items():
        worksheet.column_dimensions[letter].width = width

    # ── Row layout calculation ──
    # Material rows start at row 10 (after header, company info, buyer, column headers)
    material_end_row = MATERIAL_START_ROW + num_material_rows - 1
    spacer_row = material_end_row + 1
    total_row = spacer_row + 1
    cgst_row = total_row + 1
    sgst_row = cgst_row + 1
    roundoff_row = sgst_row + 1
    grand_total_row = roundoff_row + 1
    bank_row = grand_total_row + 1
    terms_header_row = bank_row + 1
    terms_row_2 = terms_header_row + 1
    terms_row_3 = terms_row_2 + 1

    # ── Row heights ──
    worksheet.row_dimensions[1].height = ROW_HEIGHT_COMPANY_NAME
    for row in range(2, 8):
        worksheet.row_dimensions[row].height = ROW_HEIGHT_DEFAULT
    worksheet.row_dimensions[8].height = ROW_HEIGHT_BUYER_INFO
    worksheet.row_dimensions[9].height = ROW_HEIGHT_DEFAULT

    # Spacer height shrinks as more material rows are added, keeping the
    # invoice at roughly one printed page. Minimum 10pt to avoid collapse.
    spacer_height = max(
        SPACER_MIN_HEIGHT,
        SPACER_BASE_HEIGHT - (num_material_rows - SPACER_BASELINE_ROWS) * ROW_HEIGHT_DEFAULT,
    )
    worksheet.row_dimensions[spacer_row].height = spacer_height

    for row in [total_row, cgst_row, sgst_row, roundoff_row, grand_total_row]:
        worksheet.row_dimensions[row].height = ROW_HEIGHT_DEFAULT
    worksheet.row_dimensions[bank_row].height = ROW_HEIGHT_BANK_DETAILS
    worksheet.row_dimensions[terms_header_row].height = ROW_HEIGHT_DEFAULT
    worksheet.row_dimensions[terms_row_2].height = ROW_HEIGHT_TERMS_SECOND
    worksheet.row_dimensions[terms_row_3].height = ROW_HEIGHT_TERMS_THIRD

    # ── Company header (rows 1–7) ──
    company = config["company"]
    header_lines = [
        (company["name"],                          41, False),
        (company["subtitle"],                       9, False),
        (company["address"],                        9, False),
        (f"Contact No: {company['contact']}",       9, False),
        (f"GSTN {company['gstn']}",                 9, False),
        (f"PAN NO {company['pan']}",                9, False),
        ("TAX INVOICE",                             9, True),
    ]

    for row_index, (text, font_size, is_bold) in enumerate(header_lines, start=1):
        worksheet.merge_cells(f"A{row_index}:H{row_index}")
        cell = worksheet.cell(row=row_index, column=1, value=text)
        style_cell(cell, font_size=font_size, bold=is_bold)
        _apply_merged_row_borders(worksheet, row_index)

    # ── Buyer details, invoice number, and date (row 8) ──
    buyer = config["buyer"]
    buyer_text = (
        f"{buyer['name']}\n"
        f"{buyer['address'].strip()}\n"
        f"BUYER - GSTIN {buyer['gstn']}"
    )

    worksheet.merge_cells("A8:D8")
    cell = worksheet.cell(row=8, column=1, value=buyer_text)
    style_cell(cell, bold=True)
    worksheet.cell(row=8, column=1).border = BORDER_ALL
    for column in range(2, 4):
        worksheet.cell(row=8, column=column).border = Border(
            top=Side(style="thin"), bottom=Side(style="thin"))
    worksheet.cell(row=8, column=4).border = Border(
        right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    worksheet.merge_cells("E8:F8")
    cell = worksheet.cell(row=8, column=5, value=f"INVOICE NO\n{invoice_number}-{fiscal_year}")
    style_cell(cell, bold=True)
    worksheet.cell(row=8, column=5).border = BORDER_ALL
    worksheet.cell(row=8, column=6).border = Border(
        right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    worksheet.merge_cells("G8:H8")
    cell = worksheet.cell(row=8, column=7, value=f"DATE: {invoice_date.strftime('%d-%m-%Y')}")
    style_cell(cell, bold=True)
    worksheet.cell(row=8, column=7).border = BORDER_ALL
    worksheet.cell(row=8, column=8).border = Border(
        right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    # ── Column headers (row 9) ──
    column_headers = ["Date", "Material", "Site", "Challan", "Quantity", "Rate", "Per", "Amount"]
    for column_index, header_text in enumerate(column_headers, start=1):
        cell = worksheet.cell(row=9, column=column_index, value=header_text)
        style_cell(cell, bold=True, border=BORDER_ALL)

    # ── Material data rows ──
    current_row = MATERIAL_START_ROW

    for site_info in sites_data:
        site_name = site_info["site_name"]
        date_label = site_info["date_range_label"]
        material_data = site_info["material_data"]

        first_row = current_row
        last_row = current_row + len(materials) - 1

        # Merge the date and challan columns across all materials at this site
        if last_row > first_row:
            worksheet.merge_cells(f"A{first_row}:A{last_row}")
            worksheet.merge_cells(f"D{first_row}:D{last_row}")

        cell = worksheet.cell(row=first_row, column=1, value=date_label)
        style_cell(cell, font_size=10, border=BORDER_SIDES_TOP)

        cell = worksheet.cell(row=first_row, column=4, value="As Per List")
        style_cell(cell, border=BORDER_SIDES_TOP)

        for material_index, material_name in enumerate(materials):
            row = current_row + material_index
            material_info = material_data.get(material_name, {})
            quantity = material_info.get("qty")
            rate = material_info.get("rate")
            is_first_material = (material_index == 0)

            # First row of each site gets a top border, subsequent rows get side-only
            border_style = BORDER_SIDES_TOP if is_first_material else BORDER_SIDES

            # Material name
            worksheet.cell(row=row, column=2, value=material_name)
            style_cell(worksheet.cell(row=row, column=2), border=border_style)

            # Site name
            worksheet.cell(row=row, column=3, value=site_name)
            style_cell(worksheet.cell(row=row, column=3), border=border_style)

            # Quantity (blank if this material doesn't exist at this site)
            if quantity is not None:
                worksheet.cell(row=row, column=5, value=round(quantity, 2))
            style_cell(worksheet.cell(row=row, column=5), border=border_style)

            # Rate (per-material, not per-site)
            if rate is not None:
                worksheet.cell(row=row, column=6, value=float(rate))
            style_cell(worksheet.cell(row=row, column=6), border=border_style)

            # Unit
            worksheet.cell(row=row, column=7, value=unit)
            style_cell(worksheet.cell(row=row, column=7), border=border_style)

            # Amount = Quantity × Rate (blank if material absent at this site)
            if quantity is not None and rate is not None:
                amount = round(quantity, 2) * float(rate)
            else:
                amount = None
            worksheet.cell(row=row, column=8, value=amount)
            style_cell(worksheet.cell(row=row, column=8), font_size=10, border=border_style)

            # Side borders on merged cells (date and challan) for non-first rows
            if material_index > 0:
                worksheet.cell(row=row, column=1).border = BORDER_SIDES
                worksheet.cell(row=row, column=4).border = BORDER_SIDES

        current_row = last_row + 1

    # ── Spacer row (fills remaining page height) ──
    _apply_side_borders(worksheet, spacer_row)

    # ── TOTAL / CGST / SGST rows ──
    _write_tax_line(worksheet, total_row, "TOTAL", rate=None, amount=data["total_amount"])
    _write_tax_line(worksheet, cgst_row, "CGST", rate=config["gst"]["cgst_rate"],
                    amount=data["cgst_amount"])
    _write_tax_line(worksheet, sgst_row, "SGST", rate=config["gst"]["sgst_rate"],
                    amount=data["sgst_amount"])

    # HSN code goes on the SGST row
    cell = worksheet.cell(row=sgst_row, column=2, value="HSN Code")
    style_cell(cell, font_size=10, border=BORDER_SIDES)
    cell = worksheet.cell(row=sgst_row, column=3, value=float(config["gst"]["hsn_code"]))
    style_cell(cell, font_size=10, border=BORDER_SIDES)

    # ── Round-off row ──
    _apply_side_borders(worksheet, roundoff_row, end_col=TOTAL_COLUMNS - 1,
                         border=BORDER_SIDES_BOTTOM)
    cell = worksheet.cell(row=roundoff_row, column=5, value="ROUND OFF")
    style_cell(cell, border=BORDER_SIDES_BOTTOM)
    cell = worksheet.cell(row=roundoff_row, column=8, value=data["round_off"])
    style_cell(cell, font_size=10, number_format=CURRENCY_FORMAT, border=BORDER_SIDES_BOTTOM)

    # ── Grand total row ──
    worksheet.merge_cells(f"A{grand_total_row}:G{grand_total_row}")
    cell = worksheet.cell(row=grand_total_row, column=1, value="GRAND TOTAL")
    style_cell(cell, bold=True, border=BORDER_ALL)
    for column in range(2, 8):
        worksheet.cell(row=grand_total_row, column=column).border = Border(
            top=Side(style="thin"), bottom=Side(style="thin"),
            right=Side(style="thin") if column == 7 else None,
        )
    cell = worksheet.cell(row=grand_total_row, column=8, value=data["grand_total"])
    style_cell(cell, font_size=10, number_format=CURRENCY_FORMAT, border=BORDER_ALL)

    # ── Bank details and signature (row below grand total) ──
    bank = config["bank"]
    bank_text = (
        f"BANK DETAILS :- {bank['account_name']} "
        f"BANK NAME: {bank['bank_name']}\n"
        f"A/C NO: {bank['account_no']}\n"
        f"BRANCH & IFSC : {bank['branch']} / {bank['ifsc']}"
    )
    worksheet.merge_cells(f"A{bank_row}:E{bank_row}")
    cell = worksheet.cell(row=bank_row, column=1, value=bank_text)
    style_cell(cell, bold=True)
    worksheet.cell(row=bank_row, column=1).border = BORDER_ALL
    for column in range(2, 5):
        worksheet.cell(row=bank_row, column=column).border = Border(
            top=Side(style="thin"), bottom=Side(style="thin"))
    worksheet.cell(row=bank_row, column=5).border = Border(
        right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    worksheet.merge_cells(f"F{bank_row}:H{bank_row}")
    cell = worksheet.cell(row=bank_row, column=6, value=company["name"])
    style_cell(cell, font_size=10, bold=True)
    worksheet.cell(row=bank_row, column=6).border = BORDER_ALL
    for column in range(7, 9):
        worksheet.cell(row=bank_row, column=column).border = Border(
            top=Side(style="thin"), bottom=Side(style="thin"),
            right=Side(style="thin") if column == 8 else None,
        )

    # ── Terms & Conditions ──
    worksheet.merge_cells(f"A{terms_header_row}:B{terms_header_row}")
    worksheet.merge_cells(f"C{terms_header_row}:H{terms_header_row}")
    cell = worksheet.cell(row=terms_header_row, column=1, value="Terms & Conditions :")
    style_cell(cell)
    for column in range(1, 9):
        worksheet.cell(row=terms_header_row, column=column).border = BORDER_ALL

    terms_content = [
        (terms_row_2, "1) interest @ 18% will be charged on payment due for more than "
                       "30 days from date of invoice.", 7),
        (terms_row_3, "2) Our responsibility ceases if any error is not reported in "
                       "writing 7 days from date of invoice.", 7),
        (terms_row_3 + 1,
         '"Input tax Credit of CGST / SGST / IGST charged on goods and services '
         "used exclusively or partly in supplying goods transport\n"
         'agency services has not been taken."', 10),
    ]

    for row_number, text, font_size in terms_content:
        worksheet.merge_cells(f"A{row_number}:H{row_number}")
        cell = worksheet.cell(row=row_number, column=1, value=text)
        style_cell(cell, font_size=font_size)
        for column in range(1, 9):
            worksheet.cell(row=row_number, column=column).border = BORDER_ALL


def _create_list_sheet(workbook: Workbook, data: dict) -> None:
    """
    Create the detail List sheet: a flat table of every delivery record.

    Each row shows one delivery with Date, Challan, Vehicle, Site, Material,
    Quantity, Rate, Per, and the computed Amount (Quantity × Rate).
    """
    invoice_number = data["inv_num"]
    dataframe = data["df"]

    sheet_name = sanitize_sheet_name(f"List - {invoice_number}")
    worksheet = workbook.create_sheet(title=sheet_name)

    # ── Column headers ──
    headers = ["Date", "Challan No.", "Vehicle No.", "Site", "Material",
               "Quantity", "Rate", "Per", "Amount"]
    for column_index, header_text in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=header_text)
        cell.font = Font(name="Arial", bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER_ALL

    # ── Data rows ──
    for row_offset, (_, row_data) in enumerate(dataframe.iterrows()):
        row_number = row_offset + 2

        # Safe conversions: Challan/Vehicle may be NaN
        challan = float(row_data["Challan No."]) if pd.notna(row_data["Challan No."]) else None
        vehicle = float(row_data["Vehicle No."]) if pd.notna(row_data["Vehicle No."]) else None
        quantity = float(row_data["Quantity"]) if pd.notna(row_data["Quantity"]) else 0.0
        rate = float(row_data["Rate"]) if pd.notna(row_data["Rate"]) else 0.0
        per_unit = row_data["Per"] if pd.notna(row_data.get("Per", None)) else ""
        amount = quantity * rate

        values = [row_data["Date"], challan, vehicle, row_data["Site"],
                  row_data["Material"], quantity, rate, per_unit, amount]

        for column_index, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row_number, column=column_index, value=value)
            cell.font = Font(name="Arial")
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER_SIDES
            if column_index == 9:
                cell.number_format = NUMBER_FORMAT

    # ── Totals row ──
    totals_row = len(dataframe) + 2

    quantity_total = dataframe["Quantity"].astype(float).sum()
    cell = worksheet.cell(row=totals_row, column=6, value=quantity_total)
    cell.font = Font(name="Arial")
    cell.alignment = Alignment(horizontal="center")
    cell.border = BORDER_ALL
    cell.number_format = NUMBER_FORMAT

    amount_total = round(
        (dataframe["Quantity"].astype(float) * dataframe["Rate"].astype(float)).sum(), 2
    )
    cell = worksheet.cell(row=totals_row, column=9, value=amount_total)
    cell.font = Font(name="Arial")
    cell.alignment = Alignment(horizontal="center")
    cell.border = BORDER_ALL
    cell.number_format = NUMBER_FORMAT


# ─── PDF Generation ──────────────────────────────────────────────────────────

class InvoicePDF(FPDF):
    """PDF invoice renderer with automatic header and footer."""

    def __init__(self, config: dict):
        super().__init__()
        self.config = config
        self.set_auto_page_break(auto=True, margin=15)

    def header(self) -> None:
        """Render the company header at the top of each page."""
        company = self.config["company"]

        self.set_font("Helvetica", "B", 18)
        self.cell(0, 10, company["name"], align="C", new_x="LMARGIN", new_y="NEXT")

        self.set_font("Helvetica", "", 9)
        self.cell(0, 5, company["subtitle"], align="C", new_x="LMARGIN", new_y="NEXT")
        self.cell(0, 5, company["address"], align="C", new_x="LMARGIN", new_y="NEXT")
        self.cell(0, 5, f"Contact: {company['contact']}", align="C",
                  new_x="LMARGIN", new_y="NEXT")
        self.cell(0, 5, f"GSTN: {company['gstn']}  |  PAN: {company['pan']}",
                  align="C", new_x="LMARGIN", new_y="NEXT")

        self.set_font("Helvetica", "B", 11)
        self.cell(0, 8, "TAX INVOICE", align="C", new_x="LMARGIN", new_y="NEXT", border="TB")
        self.ln(2)

    def footer(self) -> None:
        """Render the version footer at the bottom of each page."""
        self.set_y(-15)
        self.set_font("Helvetica", "I", 7)
        self.cell(0, 10, f"Generated by Invoice Generator v{__version__}", align="C")


# Column widths for the PDF table (must sum to ~190mm for A4 page width)
PDF_COLUMN_WIDTHS = [22, 28, 22, 18, 20, 20, 16, 22]


def _render_pdf(data: dict) -> InvoicePDF:
    """
    Build a PDF representation of the invoice from processed data.

    The PDF mirrors the Excel Bill sheet's content in a print-friendly format:
    invoice metadata → buyer details → itemized table → tax summary → bank details.
    """
    config = data["config"]
    invoice_number = data["inv_num"]
    fiscal_year = data["fiscal_year"]
    invoice_date = data["invoice_date"]
    sites_data = data["sites_data"]
    materials = data["materials"]
    unit = data["unit"]
    buyer = config["buyer"]
    bank = config["bank"]

    pdf = InvoicePDF(config)
    pdf.add_page()

    # ── Invoice metadata ──
    pdf.set_font("Helvetica", "B", 9)
    half_width = 95
    pdf.cell(half_width, 6, f"Invoice No: {invoice_number}-{fiscal_year}", border=1)
    pdf.cell(half_width, 6, f"Date: {invoice_date.strftime('%d-%m-%Y')}",
             border=1, new_x="LMARGIN", new_y="NEXT")

    # ── Buyer details ──
    pdf.set_font("Helvetica", "", 8)
    buyer_block = f"To: {buyer['name']}\n{buyer['address'].strip()}\nGSTIN: {buyer['gstn']}"
    pdf.multi_cell(0, 4, buyer_block, border=1)
    pdf.ln(3)

    # ── Table header ──
    column_headers = ["Date", "Material", "Site", "Challan", "Qty", "Rate", "Per", "Amount"]
    pdf.set_font("Helvetica", "B", 8)
    for header_index, header_text in enumerate(column_headers):
        pdf.cell(PDF_COLUMN_WIDTHS[header_index], 6, header_text, border=1, align="C")
    pdf.ln()

    # ── Table body ──
    pdf.set_font("Helvetica", "", 7.5)
    for site_info in sites_data:
        date_text = site_info["date_range_label"].replace("\n", " ")

        for material_index, material_name in enumerate(materials):
            material_info = site_info["material_data"].get(material_name, {})
            quantity = material_info.get("qty")
            rate = material_info.get("rate")

            if quantity is not None and rate is not None:
                amount_text = f"{round(quantity, 2) * float(rate):,.2f}"
            else:
                amount_text = ""

            row_values = [
                date_text if material_index == 0 else "",
                material_name,
                site_info["site_name"],
                "As Per List" if material_index == 0 else "",
                f"{quantity:.2f}" if quantity else "",
                f"{rate:.0f}" if rate else "",
                unit,
                amount_text,
            ]

            for col_index, value in enumerate(row_values):
                pdf.cell(PDF_COLUMN_WIDTHS[col_index], 5, str(value), border="LR", align="C")
            pdf.ln()

    # ── Separator ──
    pdf.cell(sum(PDF_COLUMN_WIDTHS), 0, "", border="T")
    pdf.ln()

    # ── Tax summary ──
    label_width = sum(PDF_COLUMN_WIDTHS[:7])
    amount_width = PDF_COLUMN_WIDTHS[7]

    def write_summary_line(label: str, value: float, bold: bool = False) -> None:
        pdf.set_font("Helvetica", "B" if bold else "", 8)
        pdf.cell(label_width, 6, label, border="LR", align="R")
        pdf.cell(amount_width, 6, f"{value:,.2f}", border="LR", align="C")
        pdf.ln()

    cgst_pct = int(config["gst"]["cgst_rate"] * 100)
    sgst_pct = int(config["gst"]["sgst_rate"] * 100)

    write_summary_line("TOTAL", data["total_amount"], bold=True)
    write_summary_line(f"CGST @ {cgst_pct}%", data["cgst_amount"])
    write_summary_line(f"SGST @ {sgst_pct}%", data["sgst_amount"])
    write_summary_line("ROUND OFF", data["round_off"])

    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(label_width, 7, "GRAND TOTAL", border=1, align="R")
    pdf.cell(amount_width, 7, f"Rs. {data['grand_total']:,}", border=1, align="C")
    pdf.ln()

    # ── HSN code ──
    pdf.set_font("Helvetica", "", 7)
    pdf.cell(0, 5, f"HSN Code: {config['gst']['hsn_code']}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    # ── Bank details ──
    pdf.set_font("Helvetica", "B", 8)
    pdf.cell(0, 5, "Bank Details:", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 8)
    pdf.cell(0, 5,
             f"{bank['account_name']} | {bank['bank_name']} | "
             f"A/C: {bank['account_no']} | Branch: {bank['branch']} | "
             f"IFSC: {bank['ifsc']}",
             new_x="LMARGIN", new_y="NEXT")

    return pdf


# ─── Public API ──────────────────────────────────────────────────────────────

def generate_invoice(
    input_file: Union[str, Path],
    inv_num: int,
    config: Optional[dict] = None,
) -> Workbook:
    """
    Generate a GST tax invoice as an Excel workbook.

    Args:
        input_file: Path to the input Excel data (.xlsx or .xls).
        inv_num: Invoice number to assign.
        config: Business configuration dict. Loads from config.yaml if not provided.

    Returns:
        openpyxl.Workbook containing a Bill sheet and a List sheet.
    """
    data = process_input_data(input_file, inv_num, config)
    workbook = Workbook()
    workbook.remove(workbook.active)
    _create_invoice_sheet(workbook, data)
    _create_list_sheet(workbook, data)
    return workbook


def generate_pdf(
    input_file: Union[str, Path],
    inv_num: int,
    output_path: Optional[Union[str, Path]] = None,
    config: Optional[dict] = None,
) -> Union[bytes, Path]:
    """
    Generate a GST tax invoice as a PDF.

    Args:
        input_file: Path to the input Excel data.
        inv_num: Invoice number to assign.
        output_path: Where to save the PDF. If None, returns PDF as bytes.
        config: Business configuration dict.

    Returns:
        Path to the saved PDF file, or raw PDF bytes if output_path is None.
    """
    data = process_input_data(input_file, inv_num, config)
    pdf = _render_pdf(data)

    if output_path:
        pdf.output(str(output_path))
        return Path(output_path)

    return pdf.output()


def batch_process(
    input_dir: Union[str, Path],
    start_num: int,
    output_dir: Optional[Union[str, Path]] = None,
    pdf: bool = False,
    config: Optional[dict] = None,
) -> list[dict]:
    """
    Process all Excel files in a directory, generating invoices with
    auto-incrementing invoice numbers.

    Args:
        input_dir: Directory containing .xlsx/.xls files.
        start_num: Invoice number for the first file (increments from here).
        output_dir: Where to save outputs. Defaults to input_dir.
        pdf: Whether to also generate PDF for each invoice.
        config: Business configuration dict.

    Returns:
        List of result dicts, one per file:
        [{"input": str, "invoice_num": int, "excel": str, "pdf": str, "error": str}]
    """
    input_directory = Path(input_dir)
    output_directory = Path(output_dir) if output_dir else input_directory
    output_directory.mkdir(parents=True, exist_ok=True)

    if config is None:
        config = load_config()

    # Find all Excel files, excluding temporary files (prefixed with ~)
    excel_files = sorted(
        f for f in input_directory.iterdir()
        if f.suffix.lower() in (".xlsx", ".xls") and not f.name.startswith("~")
    )

    if not excel_files:
        print(f"No Excel files found in {input_directory}", file=sys.stderr)
        return []

    results = []

    for file_index, file_path in enumerate(excel_files):
        invoice_number = start_num + file_index
        result = {
            "input": str(file_path),
            "invoice_num": invoice_number,
            "excel": None,
            "pdf": None,
            "error": None,
        }

        try:
            # Generate Excel
            workbook = generate_invoice(file_path, invoice_number, config)
            excel_output = output_directory / f"Invoice_{invoice_number}.xlsx"
            workbook.save(str(excel_output))
            result["excel"] = str(excel_output)
            print(f"  ✅ [{invoice_number}] {file_path.name} → {excel_output.name}")

            # Generate PDF (if requested)
            if pdf:
                pdf_output = output_directory / f"Invoice_{invoice_number}.pdf"
                generate_pdf(file_path, invoice_number, pdf_output, config)
                result["pdf"] = str(pdf_output)
                print(f"        📄 {pdf_output.name}")

        except Exception as error:
            result["error"] = str(error)
            print(f"  ❌ [{invoice_number}] {file_path.name}: {error}", file=sys.stderr)

        results.append(result)

    return results


# ─── CLI ─────────────────────────────────────────────────────────────────────

def main() -> None:
    """Command-line interface entry point."""
    parser = argparse.ArgumentParser(
        description="Invoice Generator — GST Tax Invoice from Raw Delivery Data",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  %(prog)s                              # Interactive mode\n"
            "  %(prog)s -i 178                        # Single invoice #178\n"
            "  %(prog)s -i 178 --pdf                  # Excel + PDF\n"
            "  %(prog)s --batch ./data/ --start 178   # Batch process folder\n"
        ),
    )
    parser.add_argument("-i", "--invoice", type=int, help="Invoice number")
    parser.add_argument("--input", type=str, default=str(INPUT_FILE),
                        help=f"Input Excel file (default: {INPUT_FILE})")
    parser.add_argument("--output", type=str, default=str(OUTPUT_EXCEL),
                        help=f"Output Excel file (default: {OUTPUT_EXCEL})")
    parser.add_argument("--pdf", action="store_true",
                        help="Also generate PDF output")
    parser.add_argument("--batch", type=str, metavar="DIR",
                        help="Batch process all Excel files in a directory")
    parser.add_argument("--start", type=int, default=1,
                        help="Starting invoice number for batch mode (default: 1)")
    parser.add_argument("--output-dir", type=str, metavar="DIR",
                        help="Output directory for batch mode")
    parser.add_argument("--config", type=str,
                        help="Path to config YAML file")
    parser.add_argument("--version", action="version",
                        version=f"%(prog)s {__version__}")

    args = parser.parse_args()
    config = load_config(args.config)

    # ── Batch mode ──
    if args.batch:
        batch_dir = Path(args.batch)
        if not batch_dir.is_dir():
            print(f"Error: {args.batch} is not a directory.", file=sys.stderr)
            sys.exit(1)

        print(f"📦 Batch processing: {batch_dir}/ (starting at #{args.start})")
        results = batch_process(batch_dir, args.start, args.output_dir, args.pdf, config)

        succeeded = sum(1 for r in results if r["error"] is None)
        failed = len(results) - succeeded
        print(f"\n✅ {succeeded} succeeded, ❌ {failed} failed")
        sys.exit(1 if failed > 0 else 0)

    # ── Single-file mode ──
    invoice_number = args.invoice
    if invoice_number is None:
        try:
            invoice_number = int(input("Enter invoice number: "))
        except (ValueError, EOFError):
            print("Error: Please enter a valid integer.", file=sys.stderr)
            sys.exit(1)

    try:
        workbook = generate_invoice(args.input, invoice_number, config)
        workbook.save(args.output)
        print(f"✅ Excel: {args.output}")

        if args.pdf:
            pdf_path = Path(args.output).with_suffix(".pdf")
            generate_pdf(args.input, invoice_number, pdf_path, config)
            print(f"📄 PDF:   {pdf_path}")

    except (FileNotFoundError, ValueError) as error:
        print(f"Error: {error}", file=sys.stderr)
        sys.exit(1)
    except Exception as error:
        print(f"Unexpected error: {error}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
