"""
Test suite for Invoice Generator v0.2.

Covers: data processing, edge cases, failure scenarios,
        Excel generation, PDF generation, batch processing, and config loading.
"""

import os
import tempfile
from datetime import datetime
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

# Add project root to path
import sys
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from invoice_generator.generator import (
    __version__,
    batch_process,
    format_date_short,
    generate_invoice,
    generate_pdf,
    get_fiscal_year,
    get_invoice_date,
    load_config,
    parse_dates,
    process_input_data,
    round_conventional,
    sanitize_sheet_name,
)


# ─── Fixtures ────────────────────────────────────────────────────────────────

@pytest.fixture
def sample_excel(tmp_path):
    """Create a simple valid Excel file for testing."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Challan No.", "Vehicle No.", "Site", "Material", "Quantity", "Rate", "Per"])
    ws.append([datetime(2025, 2, 10), 101, 5001, "SiteA", "Sand", 50.0, 400, "Tonne"])
    ws.append([datetime(2025, 2, 15), 102, 5002, "SiteA", "Gravel", 30.0, 350, "Tonne"])
    ws.append([datetime(2025, 2, 20), 103, 5003, "SiteB", "Sand", 40.0, 420, "Tonne"])
    path = tmp_path / "test_data.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def single_row_excel(tmp_path):
    """Minimal: 1 row of data."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Challan No.", "Vehicle No.", "Site", "Material", "Quantity", "Rate", "Per"])
    ws.append([datetime(2025, 4, 1), 301, 7001, "Thane", "Bricks", 100.0, 200, "Tonne"])
    path = tmp_path / "minimal.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def multi_rate_excel(tmp_path):
    """Different rates per material at same site."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Challan No.", "Vehicle No.", "Site", "Material", "Quantity", "Rate", "Per"])
    ws.append([datetime(2025, 3, 1), 201, 6001, "SiteA", "Sand", 40.0, 350, "Tonne"])
    ws.append([datetime(2025, 3, 5), 202, 6002, "SiteA", "Gravel", 35.0, 420, "Tonne"])
    ws.append([datetime(2025, 3, 10), 203, 6003, "SiteA", "Cement", 60.0, 500, "Tonne"])
    path = tmp_path / "multi_rate.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def string_dates_excel(tmp_path):
    """Dates as strings in d/m/Y format."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Challan No.", "Vehicle No.", "Site", "Material", "Quantity", "Rate", "Per"])
    ws.append(["15/01/2025", 401, 8001, "Powai", "TMT Bars", 20.0, 550, "Tonne"])
    ws.append(["20/01/2025", 402, 8002, "Powai", "TMT Bars", 15.5, 550, "Tonne"])
    path = tmp_path / "string_dates.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def missing_values_excel(tmp_path):
    """Has None for Challan No. and Vehicle No."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Challan No.", "Vehicle No.", "Site", "Material", "Quantity", "Rate", "Per"])
    ws.append([datetime(2025, 2, 10), None, None, "SiteA", "Sand", 50.0, 400, "Tonne"])
    ws.append([datetime(2025, 2, 15), 102, None, "SiteA", "Sand", 30.0, 400, "Tonne"])
    path = tmp_path / "missing_vals.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def sample_config():
    """Test config dict."""
    return {
        "company": {
            "name": "TEST COMPANY",
            "subtitle": "(TEST)",
            "address": "123 Test St",
            "contact": "1234567890",
            "gstn": "00XXXXX0000X0XX",
            "pan": "XXXXX0000X",
        },
        "buyer": {
            "name": "TEST BUYER",
            "address": "456 Buyer St",
            "gstn": "00YYYYY0000Y0YY",
        },
        "bank": {
            "account_name": "TEST COMPANY",
            "bank_name": "TEST BANK",
            "account_no": "000000000",
            "branch": "TEST",
            "ifsc": "TEST0000000",
        },
        "gst": {"cgst_rate": 0.09, "sgst_rate": 0.09, "hsn_code": 996511},
        "unit": "Tonne",
    }


# ─── Helper Function Tests ──────────────────────────────────────────────────

class TestFormatDateShort:
    def test_normal(self):
        assert format_date_short(datetime(2025, 2, 16)) == "16-Feb"

    def test_single_digit_day(self):
        assert format_date_short(datetime(2025, 1, 5)) == "5-Jan"

    def test_december(self):
        assert format_date_short(datetime(2025, 12, 31)) == "31-Dec"


class TestParseDates:
    def test_datetime_objects(self):
        series = pd.Series([datetime(2025, 1, 15), datetime(2025, 2, 20)])
        result = parse_dates(series)
        assert len(result) == 2

    def test_string_dates_dmy(self):
        series = pd.Series(["16/02/2025", "20/03/2025"])
        result = parse_dates(series)
        assert len(result) == 2
        assert result[0].day == 16
        assert result[0].month == 2

    def test_string_dates_iso(self):
        series = pd.Series(["2025-03-15"])
        result = parse_dates(series)
        assert len(result) == 1
        assert result[0].month == 3

    def test_mixed_formats(self):
        series = pd.Series(["16/02/2025", "2025-03-15", "05-01-2025"])
        result = parse_dates(series)
        assert len(result) == 3

    def test_skips_nan(self):
        series = pd.Series([datetime(2025, 1, 1), None, float("nan")])
        result = parse_dates(series)
        assert len(result) == 1

    def test_skips_garbage(self):
        series = pd.Series(["not-a-date", "garbage", "abc"])
        result = parse_dates(series)
        assert len(result) == 0

    def test_empty_series(self):
        result = parse_dates(pd.Series([], dtype=object))
        assert len(result) == 0

    def test_timestamps(self):
        series = pd.Series([pd.Timestamp("2025-06-15")])
        result = parse_dates(series)
        assert len(result) == 1
        assert isinstance(result[0], datetime)


class TestSanitizeSheetName:
    def test_normal(self):
        assert sanitize_sheet_name("Bill - 1") == "Bill - 1"

    def test_special_chars(self):
        assert sanitize_sheet_name("Site A/B [test]") == "Site A_B _test_"

    def test_truncation(self):
        result = sanitize_sheet_name("A" * 50)
        assert len(result) == 31

    def test_colon(self):
        assert sanitize_sheet_name("Test:Sheet") == "Test_Sheet"


class TestRoundConventional:
    def test_point_five_rounds_up(self):
        assert round_conventional(100.5) == 101

    def test_below_point_five_rounds_down(self):
        assert round_conventional(100.4) == 100

    def test_above_point_five_rounds_up(self):
        assert round_conventional(100.6) == 101

    def test_exact_integer(self):
        assert round_conventional(100.0) == 100

    def test_negative(self):
        assert round_conventional(-0.3) == 0


class TestGetFiscalYear:
    def test_q4(self):
        assert get_fiscal_year([datetime(2025, 3, 15)]) == "2024/25"

    def test_q1(self):
        assert get_fiscal_year([datetime(2025, 4, 15)]) == "2025/26"

    def test_q2(self):
        assert get_fiscal_year([datetime(2025, 9, 1)]) == "2025/26"

    def test_q3(self):
        assert get_fiscal_year([datetime(2025, 12, 31)]) == "2025/26"


class TestGetInvoiceDate:
    def test_feb(self):
        dt = get_invoice_date([datetime(2025, 2, 10)])
        assert dt == datetime(2025, 2, 28)

    def test_leap_year(self):
        dt = get_invoice_date([datetime(2024, 2, 10)])
        assert dt == datetime(2024, 2, 29)

    def test_jan(self):
        dt = get_invoice_date([datetime(2025, 1, 5)])
        assert dt == datetime(2025, 1, 31)


# ─── Config Tests ────────────────────────────────────────────────────────────

class TestLoadConfig:
    def test_loads_default_when_no_file(self, tmp_path):
        cfg = load_config(tmp_path / "nonexistent.yaml")
        assert cfg["company"]["name"] == "YOUR COMPANY NAME"

    def test_loads_yaml(self, tmp_path):
        yaml_path = tmp_path / "test_config.yaml"
        yaml_path.write_text('company:\n  name: "My Company"\n')
        cfg = load_config(yaml_path)
        assert cfg["company"]["name"] == "My Company"
        # Should still have defaults for other fields
        assert "gstn" in cfg["company"]

    def test_project_config_loads(self):
        """Test that the actual project config.yaml loads successfully."""
        cfg_path = Path(__file__).parent.parent / "config.yaml"
        if cfg_path.exists():
            cfg = load_config(cfg_path)
            assert cfg["company"]["name"] == "ACME ENTERPRISES"


# ─── Data Processing Tests ───────────────────────────────────────────────────

class TestProcessInputData:
    def test_valid_data(self, sample_excel, sample_config):
        data = process_input_data(sample_excel, 1, sample_config)
        assert data["inv_num"] == 1
        assert len(data["sites_data"]) == 2  # SiteA, SiteB
        assert len(data["materials"]) == 2  # Sand, Gravel
        assert data["total_amount"] > 0

    def test_single_row(self, single_row_excel, sample_config):
        data = process_input_data(single_row_excel, 1, sample_config)
        assert len(data["sites_data"]) == 1
        assert len(data["materials"]) == 1
        assert data["total_amount"] == 20000.0  # 100 * 200

    def test_multi_rate(self, multi_rate_excel, sample_config):
        data = process_input_data(multi_rate_excel, 1, sample_config)
        # Sand=350, Gravel=420, Cement=500 — all different rates
        mat_data = data["sites_data"][0]["material_data"]
        assert mat_data["Sand"]["rate"] == 350
        assert mat_data["Gravel"]["rate"] == 420
        assert mat_data["Cement"]["rate"] == 500

    def test_string_dates(self, string_dates_excel, sample_config):
        data = process_input_data(string_dates_excel, 1, sample_config)
        assert data["fiscal_year"] == "2024/25"

    def test_missing_challan(self, missing_values_excel, sample_config):
        data = process_input_data(missing_values_excel, 1, sample_config)
        assert data["total_amount"] == 32000.0  # (50+30) * 400

    def test_gst_calculation(self, sample_excel, sample_config):
        data = process_input_data(sample_excel, 1, sample_config)
        total = data["total_amount"]
        assert data["cgst_amount"] == round(total * 0.09, 2)
        assert data["sgst_amount"] == round(total * 0.09, 2)

    def test_grand_total_rounding(self, sample_excel, sample_config):
        data = process_input_data(sample_excel, 1, sample_config)
        gross = data["total_amount"] + data["cgst_amount"] + data["sgst_amount"]
        assert data["grand_total"] == round_conventional(gross)
        assert data["round_off"] == round(data["grand_total"] - gross, 2)

    def test_single_day_date_range(self, tmp_path, sample_config):
        """Single day should show just the date, not a range."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "Challan No.", "Vehicle No.", "Site", "Material", "Quantity", "Rate", "Per"])
        ws.append([datetime(2025, 3, 15), 101, 5001, "SiteA", "Sand", 50.0, 400, "Tonne"])
        path = tmp_path / "single_day.xlsx"
        wb.save(path)
        data = process_input_data(path, 1, sample_config)
        date_label = data["sites_data"][0]["date_range_label"]
        assert "\nto\n" not in date_label
        assert date_label == "15-Mar"


# ─── Failure Scenario Tests ─────────────────────────────────────────────────

class TestFailureScenarios:
    def test_file_not_found(self, sample_config):
        with pytest.raises(FileNotFoundError):
            process_input_data("/nonexistent/file.xlsx", 1, sample_config)

    def test_missing_columns(self, tmp_path, sample_config):
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "Site"])  # Missing most columns
        ws.append([datetime(2025, 1, 1), "SiteA"])
        path = tmp_path / "bad_cols.xlsx"
        wb.save(path)
        with pytest.raises(ValueError, match="Missing required columns"):
            process_input_data(path, 1, sample_config)

    def test_empty_data(self, tmp_path, sample_config):
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "Challan No.", "Vehicle No.", "Site", "Material", "Quantity", "Rate", "Per"])
        # No data rows
        path = tmp_path / "empty.xlsx"
        wb.save(path)
        with pytest.raises(ValueError, match="No valid data"):
            process_input_data(path, 1, sample_config)

    def test_all_nan_sites(self, tmp_path, sample_config):
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "Challan No.", "Vehicle No.", "Site", "Material", "Quantity", "Rate", "Per"])
        ws.append([datetime(2025, 1, 1), 1, 1, None, "Sand", 10, 100, "Tonne"])
        path = tmp_path / "nan_sites.xlsx"
        wb.save(path)
        with pytest.raises(ValueError, match="No valid data"):
            process_input_data(path, 1, sample_config)

    def test_non_numeric_quantity(self, tmp_path, sample_config):
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "Challan No.", "Vehicle No.", "Site", "Material", "Quantity", "Rate", "Per"])
        ws.append([datetime(2025, 1, 1), 1, 1, "SiteA", "Sand", "N/A", 100, "Tonne"])
        path = tmp_path / "bad_qty.xlsx"
        wb.save(path)
        with pytest.raises(ValueError, match="Non-numeric"):
            process_input_data(path, 1, sample_config)

    def test_non_numeric_rate(self, tmp_path, sample_config):
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "Challan No.", "Vehicle No.", "Site", "Material", "Quantity", "Rate", "Per"])
        ws.append([datetime(2025, 1, 1), 1, 1, "SiteA", "Sand", 10, "free", "Tonne"])
        path = tmp_path / "bad_rate.xlsx"
        wb.save(path)
        with pytest.raises(ValueError, match="Non-numeric"):
            process_input_data(path, 1, sample_config)

    def test_no_parseable_dates(self, tmp_path, sample_config):
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "Challan No.", "Vehicle No.", "Site", "Material", "Quantity", "Rate", "Per"])
        ws.append(["garbage", 1, 1, "SiteA", "Sand", 10, 100, "Tonne"])
        path = tmp_path / "bad_dates.xlsx"
        wb.save(path)
        with pytest.raises(ValueError, match="No valid dates"):
            process_input_data(path, 1, sample_config)


# ─── Excel Generation Tests ─────────────────────────────────────────────────

class TestGenerateInvoice:
    def test_returns_workbook(self, sample_excel, sample_config):
        wb = generate_invoice(sample_excel, 1, sample_config)
        assert isinstance(wb, Workbook)

    def test_has_two_sheets(self, sample_excel, sample_config):
        wb = generate_invoice(sample_excel, 1, sample_config)
        assert len(wb.sheetnames) == 2

    def test_sheet_names(self, sample_excel, sample_config):
        wb = generate_invoice(sample_excel, 42, sample_config)
        assert "Bill - 42" in wb.sheetnames
        assert "List - 42" in wb.sheetnames

    def test_amounts_are_numbers_not_formulas(self, sample_excel, sample_config):
        wb = generate_invoice(sample_excel, 1, sample_config)
        ws = wb["Bill - 1"]
        for r in range(10, 16):
            val = ws.cell(row=r, column=8).value
            if val is not None:
                assert isinstance(val, (int, float)), f"Row {r} Amount is {type(val)}, not number"

    def test_total_amount_correct(self, sample_excel, sample_config):
        wb = generate_invoice(sample_excel, 1, sample_config)
        ws = wb["Bill - 1"]
        # Find TOTAL row
        for r in range(10, 30):
            if ws.cell(row=r, column=5).value == "TOTAL":
                total = ws.cell(row=r, column=8).value
                # Manual: SiteA Sand 50*400=20000 + SiteA Gravel 30*350=10500
                #         SiteB Sand 40*420=16800 + SiteB Gravel (missing)=0
                assert total == 47300.0
                break

    def test_list_sheet_amounts(self, sample_excel, sample_config):
        wb = generate_invoice(sample_excel, 1, sample_config)
        ws = wb["List - 1"]
        # Row 2: qty=50, rate=400, amount=20000
        assert ws.cell(row=2, column=9).value == 20000.0
        # Row 3: qty=30, rate=350, amount=10500
        assert ws.cell(row=3, column=9).value == 10500.0

    def test_saves_to_file(self, sample_excel, sample_config, tmp_path):
        wb = generate_invoice(sample_excel, 1, sample_config)
        out = tmp_path / "output.xlsx"
        wb.save(out)
        assert out.exists()
        assert out.stat().st_size > 0

    def test_long_invoice_number(self, sample_excel, sample_config):
        """Sheet name >31 chars should be truncated."""
        wb = generate_invoice(sample_excel, 123456789012345, sample_config)
        for name in wb.sheetnames:
            assert len(name) <= 31

    def test_missing_material_shows_blank(self, sample_excel, sample_config):
        """SiteB only has Sand — Gravel row should have blank Amount."""
        wb = generate_invoice(sample_excel, 1, sample_config)
        ws = wb["Bill - 1"]
        # SiteB has 2 rows (Sand + Gravel), Gravel is missing
        # Find the SiteB Gravel row
        for r in range(10, 25):
            site = ws.cell(row=r, column=3).value
            mat = ws.cell(row=r, column=2).value
            if site == "SiteB" and mat == "Gravel":
                amt = ws.cell(row=r, column=8).value
                assert amt is None, f"Expected blank, got {amt}"
                break


# ─── PDF Generation Tests ───────────────────────────────────────────────────

class TestGeneratePDF:
    def test_returns_bytes(self, sample_excel, sample_config):
        result = generate_pdf(sample_excel, 1, config=sample_config)
        assert isinstance(result, (bytes, bytearray))
        assert len(result) > 0

    def test_saves_to_file(self, sample_excel, sample_config, tmp_path):
        out = tmp_path / "invoice.pdf"
        result = generate_pdf(sample_excel, 1, out, sample_config)
        assert out.exists()
        assert out.stat().st_size > 0

    def test_pdf_starts_with_marker(self, sample_excel, sample_config):
        pdf_bytes = generate_pdf(sample_excel, 1, config=sample_config)
        assert pdf_bytes[:5] == b"%PDF-"


# ─── Batch Processing Tests ─────────────────────────────────────────────────

class TestBatchProcess:
    def test_processes_all_files(self, tmp_path, sample_config):
        # Create 3 test files
        for i in range(3):
            wb = Workbook()
            ws = wb.active
            ws.append(["Date", "Challan No.", "Vehicle No.", "Site", "Material", "Quantity", "Rate", "Per"])
            ws.append([datetime(2025, 1, 1), 100+i, 5000+i, "Site", "Sand", 10.0+i, 100, "Tonne"])
            wb.save(tmp_path / f"data_{i}.xlsx")

        out_dir = tmp_path / "output"
        results = batch_process(tmp_path, 200, str(out_dir), pdf=False, config=sample_config)

        assert len(results) == 3
        assert all(r["error"] is None for r in results)
        assert results[0]["invoice_num"] == 200
        assert results[1]["invoice_num"] == 201
        assert results[2]["invoice_num"] == 202

    def test_batch_with_pdf(self, tmp_path, sample_config):
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "Challan No.", "Vehicle No.", "Site", "Material", "Quantity", "Rate", "Per"])
        ws.append([datetime(2025, 1, 1), 100, 5000, "Site", "Sand", 10.0, 100, "Tonne"])
        wb.save(tmp_path / "data.xlsx")

        out_dir = tmp_path / "output"
        results = batch_process(tmp_path, 1, str(out_dir), pdf=True, config=sample_config)

        assert len(results) == 1
        assert results[0]["pdf"] is not None
        assert Path(results[0]["pdf"]).exists()

    def test_batch_empty_dir(self, tmp_path, sample_config):
        empty = tmp_path / "empty"
        empty.mkdir()
        results = batch_process(str(empty), 1, config=sample_config)
        assert results == []

    def test_batch_handles_bad_files(self, tmp_path, sample_config):
        # One good file, one bad file
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "Challan No.", "Vehicle No.", "Site", "Material", "Quantity", "Rate", "Per"])
        ws.append([datetime(2025, 1, 1), 100, 5000, "Site", "Sand", 10.0, 100, "Tonne"])
        wb.save(tmp_path / "good.xlsx")

        wb2 = Workbook()
        ws2 = wb2.active
        ws2.append(["Bad", "Columns"])
        wb2.save(tmp_path / "bad.xlsx")

        results = batch_process(tmp_path, 1, config=sample_config)
        assert len(results) == 2
        good = [r for r in results if r["error"] is None]
        bad = [r for r in results if r["error"] is not None]
        assert len(good) == 1
        assert len(bad) == 1


# ─── Version Test ────────────────────────────────────────────────────────────

def test_version():
    assert __version__ == "0.2.0"
